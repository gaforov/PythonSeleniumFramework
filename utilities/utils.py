import inspect
import logging
from openpyxl import Workbook, load_workbook

class Utils:
    def assert_list_items_text(self, list_name, value):
        count = 1
        for stop in list_name:
            print("Type of flight is:", stop.text)
            assert stop.text == value
            print("(Test", count, "of", len(list_name), ") Passed!")
            count += 1

    def test_logger(logLevel=logging.DEBUG):
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        handler = logging.FileHandler("automation.log", mode="w")
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        handler.setFormatter(formatter)
        logger.addHandler(handler)
        return logger

    def read_data_from_excel_file(file_path, sheet_name):
        datalist = []
        wb = load_workbook(filename=file_path)
        sheet = wb[sheet_name]
        row_count = sheet.max_row
        column_count = sheet.max_column

        for i in range(2, row_count + 1):  # row 1 is a header, start count from 2.
            row = []
            for j in range(1, column_count + 1):
                row.append(sheet.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist